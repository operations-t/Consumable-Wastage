#!/usr/bin/env python3
"""Normalize Shwapno consumable, wastage, sales, hierarchy and target files."""

from __future__ import annotations

import csv
import datetime as dt
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


OUTLET_CODE_RE = re.compile(r"^[A-Z][0-9]{3,4}$")


def clean_text(value: Any) -> str:
    return " ".join(str(value or "").replace("\u00a0", " ").split()).strip()


def normalize_outlet_code(value: Any) -> str:
    value = clean_text(value).upper()
    return value if OUTLET_CODE_RE.fullmatch(value) else ""


def parse_date(value: Any) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = clean_text(value)
    for pattern in ("%d-%m-%Y", "%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(text, pattern).date()
        except ValueError:
            pass
    return None


def iso_date(value: dt.date | None) -> str | None:
    return value.isoformat() if value else None


def parse_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    text = clean_text(value).replace(",", "")
    if text.startswith("(") and text.endswith(")"):
        text = f"-{text[1:-1]}"
    try:
        return float(text)
    except ValueError:
        return 0.0


# Benchmark targets are lower-is-better control limits expressed in percent.
# Anything outside this band means the Target.txt column was entered in the
# wrong unit, so the build fails loudly instead of publishing a 100x error.
TARGET_MIN = 0.0001  # 0.01%
TARGET_MAX = 0.5000  # 50.00%


def parse_percent(value: Any) -> float | None:
    """Parse a benchmark target column that is documented in percent.

    "0.50%" and a bare "0.50" both mean 0.50 percent. The previous
    ``number / 100 if number > 1 else number`` heuristic silently read a bare
    ``0.5`` as 50%, so the unit is now fixed by the column, not guessed.
    """
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    if text.endswith("%"):
        text = text[:-1].strip()
    try:
        return float(text) / 100
    except ValueError:
        return None


def canonical_criteria(format_name: Any, pnp_status: Any, ownership: Any) -> str:
    pnp = clean_text(pnp_status).upper().replace(" ", "-")
    pnp = "Non-PNP" if pnp in {"NON-PNP", "NONPNP"} else "PNP"
    owner = clean_text(ownership).upper()
    owner = "Own" if owner in {"OWN", "OWNED"} else "FR" if owner == "FR" else clean_text(ownership)
    return f"{clean_text(format_name)}-{pnp}-{owner}"


def first_present(values: Iterable[str]) -> str:
    for value in values:
        if clean_text(value):
            return clean_text(value)
    return ""


TARGET_COLUMNS = {
    "consumableTarget": "Consumable % On Sales",
    "wastageSalesTarget": "Wastage % On Sales",
    "wastagePnpTarget": "Wastage % On PNP Sales",
}


def read_targets(path: Path) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    targets: dict[str, dict[str, Any]] = {}
    ordered: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter="\t")
        headers = {clean_text(name) for name in (reader.fieldnames or [])}
        required = {"Final Criteria", *TARGET_COLUMNS.values()}
        missing = sorted(required - headers)
        if missing:
            raise ValueError(
                f"{path.name} is missing columns: {', '.join(missing)}. "
                f"Found: {', '.join(sorted(headers))}."
            )
        for line_number, row in enumerate(reader, start=2):
            criteria = clean_text(row.get("Final Criteria"))
            if not criteria:
                continue
            item: dict[str, Any] = {"criteria": criteria}
            for field, column in TARGET_COLUMNS.items():
                raw = row.get(column)
                parsed = parse_percent(raw)
                if parsed is None:
                    raise ValueError(
                        f"{path.name} line {line_number}: '{column}' for "
                        f"'{criteria}' is not a percentage (found {raw!r})."
                    )
                if not TARGET_MIN <= parsed <= TARGET_MAX:
                    raise ValueError(
                        f"{path.name} line {line_number}: '{column}' for "
                        f"'{criteria}' resolves to {parsed * 100:.4f}%, outside the "
                        f"expected {TARGET_MIN * 100:g}%-{TARGET_MAX * 100:g}% band. "
                        f"Enter targets in percent, for example 0.50%."
                    )
                item[field] = parsed
            if criteria.casefold() in targets:
                raise ValueError(f"{path.name}: duplicate Final Criteria '{criteria}'.")
            targets[criteria.casefold()] = item
            ordered.append(item)
    if not ordered:
        raise ValueError(f"{path.name} contains no benchmark criteria rows.")
    return targets, ordered


def read_zone_distribution(path: Path, targets: dict[str, dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "Final_Zone Dis" if "Final_Zone Dis" in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    index = {header: position for position, header in enumerate(headers)}
    required = {"CODE", "Outlet Name", "Format", "Division", "District", "PNP Non PNP status", "Status"}
    missing = sorted(required - set(index))
    if missing:
        raise ValueError(f"Zone distribution is missing columns: {', '.join(missing)}")

    def value(row: tuple[Any, ...], name: str) -> Any:
        position = index.get(name)
        return row[position] if position is not None and position < len(row) else None

    records: dict[str, dict[str, Any]] = {}
    duplicates: list[str] = []
    unknown_criteria: Counter[str] = Counter()
    row_count = 0
    for row in rows:
        code = normalize_outlet_code(value(row, "CODE"))
        if not code:
            continue
        row_count += 1
        if code in records:
            duplicates.append(code)
        criteria = canonical_criteria(
            value(row, "Format"), value(row, "PNP Non PNP status"), value(row, "Status")
        )
        target = targets.get(criteria.casefold())
        if target is None:
            unknown_criteria[criteria] += 1
        launch_date = parse_date(value(row, "Launching Date"))
        records[code] = {
            "code": code,
            "name": clean_text(value(row, "Outlet Name")) or code,
            "regionalLeader": first_present(
                [value(row, "Leader"), value(row, "Regional Head HR Name")]
            ),
            "regionalLeaderFullName": clean_text(value(row, "Regional Head HR Name")),
            "zone": first_present([value(row, "Zonal"), value(row, "Zonal HR Name")]),
            "zonalFullName": clean_text(value(row, "Zonal HR Name")),
            "division": clean_text(value(row, "Division")),
            "district": clean_text(value(row, "District")),
            "area": clean_text(value(row, "Area")),
            "format": clean_text(value(row, "Format")),
            "pnpStatus": "PNP" if "NON" not in clean_text(value(row, "PNP Non PNP status")).upper() else "Non-PNP",
            "ownership": "Own" if clean_text(value(row, "Status")).upper() == "OWN" else clean_text(value(row, "Status")),
            "locationType": clean_text(value(row, "Location Type")),
            "locationClass": clean_text(value(row, "Location Type(Dv,Ds,T)")),
            "populationDensity": clean_text(value(row, "Population Density")),
            "incomeLevel": clean_text(value(row, "Income level")),
            "launchDate": iso_date(launch_date),
            "sft": parse_number(value(row, "SFT")) or None,
            "criteria": target["criteria"] if target else criteria,
            "consumableTarget": target["consumableTarget"] if target else None,
            "wastageSalesTarget": target["wastageSalesTarget"] if target else None,
            "wastagePnpTarget": target["wastagePnpTarget"] if target else None,
            "mapped": True,
        }
    workbook.close()
    quality = {
        "sheet": sheet_name,
        "rows": row_count,
        "duplicateCodes": sorted(set(duplicates)),
        "unknownCriteria": dict(unknown_criteria),
    }
    return records, quality


def read_sales(path: Path) -> tuple[dict[tuple[str, dt.date], dict[str, float]], dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "Comparative" if "Comparative" in workbook.sheetnames else workbook.sheetnames[0]
    sheet = workbook[sheet_name]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    index = {header: position for position, header in enumerate(headers)}
    required = {"Outlet Code", "Date", "Division", "POS NSI"}
    missing = sorted(required - set(index))
    if missing:
        raise ValueError(f"Sales file is missing columns: {', '.join(missing)}")

    def value(row: tuple[Any, ...], name: str) -> Any:
        position = index.get(name)
        return row[position] if position is not None and position < len(row) else None

    daily: dict[tuple[str, dt.date], dict[str, float]] = defaultdict(
        lambda: {"sales": 0.0, "pnpSales": 0.0}
    )
    division_rows: Counter[str] = Counter()
    division_value: dict[str, float] = defaultdict(float)
    dates: list[dt.date] = []
    invalid_rows = 0
    row_count = 0
    detail_total = 0.0
    reported_total: float | None = None
    for row in rows:
        raw_code = clean_text(value(row, "Outlet Code"))
        sales_value = parse_number(value(row, "POS NSI"))
        # The export ends with a "Total (999)" footer row. Capture it so the
        # detail rows can be reconciled against the source's own total.
        if raw_code.upper().startswith("TOTAL"):
            reported_total = sales_value
            continue
        code = normalize_outlet_code(raw_code)
        date = parse_date(value(row, "Date"))
        if not code or date is None:
            if raw_code or sales_value:
                invalid_rows += 1
            continue
        division = clean_text(value(row, "Division"))
        daily[(code, date)]["sales"] += sales_value
        # User-confirmed business rule: FRESH PRODUCE is PNP Sales.
        if division.upper() == "FRESH PRODUCE":
            daily[(code, date)]["pnpSales"] += sales_value
        division_rows[division] += 1
        division_value[division] += sales_value
        detail_total += sales_value
        dates.append(date)
        row_count += 1
    workbook.close()

    reconciliation: dict[str, Any] = {
        "detailTotal": round_money(detail_total),
        "reportedTotal": None if reported_total is None else round_money(reported_total),
        "difference": None,
        "differenceRatio": None,
        "matches": None,
    }
    if reported_total is not None:
        difference = detail_total - reported_total
        reconciliation["difference"] = round_money(difference)
        reconciliation["differenceRatio"] = (
            abs(difference) / abs(reported_total) if reported_total else None
        )
        # Tight enough to surface a real export gap, loose enough to ignore
        # currency rounding. Anything above 0.1% blocks the build in
        # validate_payload; below that it is reported, not fatal.
        reconciliation["matches"] = abs(difference) <= max(1.0, abs(reported_total) * 1e-7)

    quality = {
        "sheet": sheet_name,
        "rows": row_count,
        "invalidRows": invalid_rows,
        "dateMin": iso_date(min(dates) if dates else None),
        "dateMax": iso_date(max(dates) if dates else None),
        "outlets": len({code for code, _ in daily}),
        "divisionRows": dict(sorted(division_rows.items())),
        "divisionValue": {name: round_money(total) for name, total in sorted(division_value.items())},
        "reconciliation": reconciliation,
    }
    return daily, quality


# SAP ALV exports write a variable-length selection-criteria preamble before the
# column header, and the header labels differ slightly between layouts (the
# consumable export says "Quantity in UnE" where the wastage export says
# "Qty in UnE"). Locate columns by name so a layout change fails loudly instead
# of quietly reading the wrong column.
SAP_COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "postingDate": ("Pstng Date", "Posting Date", "Pstng date", "Post Date"),
    "plant": ("Plnt", "Plant"),
    "amount": ("Amount in LC", "Amount in local currency", "Amount LC"),
    "name": ("Name 1", "Name1", "Name"),
    "material": ("Material",),
    "materialDescription": ("Material Description", "Material description", "Description"),
    "quantity": ("Quantity in UnE", "Qty in UnE", "Quantity", "Qty"),
    "unit": ("EUn", "BUn", "UoM"),
    "movement": ("MvT", "Movement type", "MvT.", "Mvt"),
    "documentDate": ("Doc. Date", "Doc Date", "Document Date"),
}
SAP_REQUIRED_COLUMNS = ("postingDate", "plant", "amount")
SAP_HEADER_SEARCH_ROWS = 200


def locate_sap_header(cells: list[str]) -> dict[str, int] | None:
    """Return a field -> column index map if this row is the ALV header."""
    if len(cells) < 10:
        return None
    lookup = {clean_text(cell).casefold(): position for position, cell in enumerate(cells) if clean_text(cell)}
    mapping: dict[str, int] = {}
    for field, aliases in SAP_COLUMN_ALIASES.items():
        for alias in aliases:
            position = lookup.get(alias.casefold())
            if position is not None:
                mapping[field] = position
                break
    if any(field not in mapping for field in SAP_REQUIRED_COLUMNS):
        return None
    return mapping


def read_sap_text_export(
    path: Path,
) -> tuple[dict[tuple[str, dt.date], float], dict[str, str], dict[str, Any], dict[str, Any]]:
    daily: dict[tuple[str, dt.date], float] = defaultdict(float)
    outlet_names: dict[str, str] = {}
    # Material detail was previously read and discarded. Keeping it turns
    # "this outlet wastes too much" into "this outlet wastes these items".
    material_labels: dict[str, tuple[str, str]] = {}
    material_by_outlet: dict[tuple[str, str], list[float]] = defaultdict(lambda: [0.0, 0.0])
    material_totals: dict[str, list[float]] = defaultdict(lambda: [0.0, 0.0])
    movement_rows: Counter[str] = Counter()
    movement_value: dict[str, float] = defaultdict(float)
    material_codes: set[str] = set()
    dates: list[dt.date] = []
    positive_amount_rows = 0
    negative_amount_rows = 0
    row_count = 0
    skipped_rows = 0
    repeated_headers = 0
    columns: dict[str, int] | None = None
    header_row = None

    with path.open("r", encoding="utf-16", newline="") as handle:
        for line_number, row in enumerate(csv.reader(handle, delimiter="\t")):
            cells = [clean_text(cell) for cell in row]
            if columns is None:
                if line_number > SAP_HEADER_SEARCH_ROWS:
                    break
                found = locate_sap_header(cells)
                if found is not None:
                    columns = found
                    header_row = line_number + 1
                continue

            width = max(columns.values()) + 1
            if len(cells) < width:
                continue

            def field(name: str) -> str:
                position = columns.get(name) if columns else None
                return cells[position] if position is not None else ""

            code = normalize_outlet_code(field("plant"))
            # Posting Date is the operating period; Document Date can fall before it.
            posting_date = parse_date(field("postingDate"))
            if not code or posting_date is None:
                # SAP repeats the column header on every printed page. Those are
                # expected structure, not dropped data, so keep them out of the
                # skipped count that signals a parsing problem.
                if any(cells) and locate_sap_header(cells) is None:
                    skipped_rows += 1
                else:
                    repeated_headers += int(locate_sap_header(cells) is not None)
                continue

            amount = parse_number(field("amount"))
            movement = field("movement")
            # SAP issues are negative and reversals are positive.
            # Negating the sum gives net usage / net wastage.
            daily[(code, posting_date)] += -amount
            movement_value[movement] += -amount
            name = field("name")
            if name:
                outlet_names[code] = name
            movement_rows[movement] += 1

            material = field("material")
            material_codes.add(material)
            if material:
                quantity = -parse_number(field("quantity"))
                if material not in material_labels:
                    material_labels[material] = (field("materialDescription"), field("unit"))
                bucket = material_by_outlet[(code, material)]
                bucket[0] += -amount
                bucket[1] += quantity
                total = material_totals[material]
                total[0] += -amount
                total[1] += quantity

            dates.append(posting_date)
            row_count += 1
            positive_amount_rows += int(amount > 0)
            negative_amount_rows += int(amount < 0)

    if columns is None:
        raise ValueError(
            f"{path.name}: could not find the SAP column header within the first "
            f"{SAP_HEADER_SEARCH_ROWS} rows. Expected labels including "
            f"'Pstng Date', 'Plnt' and 'Amount in LC'. Re-export with the standard "
            f"layout and keep the Unicode (UTF-16) tab-separated format."
        )
    if row_count == 0:
        raise ValueError(
            f"{path.name}: the column header was found on row {header_row} but no "
            f"movement rows parsed. Check that the Plnt column holds outlet codes."
        )

    quality = {
        "rows": row_count,
        "headerRow": header_row,
        "columns": {name: position for name, position in sorted(columns.items())},
        "skippedRows": skipped_rows,
        "repeatedHeaderRows": repeated_headers,
        "dateMin": iso_date(min(dates) if dates else None),
        "dateMax": iso_date(max(dates) if dates else None),
        "outlets": len({code for code, _ in daily}),
        "materials": len(material_codes),
        "movementRows": dict(sorted(movement_rows.items())),
        "movementValue": {name: round_money(value) for name, value in sorted(movement_value.items())},
        "issueRows": negative_amount_rows,
        "reversalRows": positive_amount_rows,
    }
    detail = {
        "labels": material_labels,
        "byOutlet": material_by_outlet,
        "totals": material_totals,
    }
    return daily, outlet_names, quality, detail


TOP_MATERIALS_PER_OUTLET = 6
TOP_MATERIALS_OVERALL = 25


def build_material_view(
    consumable_detail: dict[str, Any],
    wastage_detail: dict[str, Any],
) -> tuple[list[list[Any]], dict[str, dict[str, list[list[float]]]], dict[str, list[list[Any]]]]:
    """Compact the material detail into a shared catalog plus per-outlet leaders.

    Every material appears once in ``catalog``; outlet rows reference it by
    index, which keeps the payload small enough to ship to a browser.
    """
    catalog_index: dict[str, int] = {}
    catalog: list[list[Any]] = []

    def material_id(code: str, source: dict[str, Any]) -> int:
        if code not in catalog_index:
            description, unit = source["labels"].get(code, ("", ""))
            catalog_index[code] = len(catalog)
            catalog.append([code, description or code, unit])
        return catalog_index[code]

    def top_rows(pairs: list[tuple[str, list[float]]], source: dict[str, Any], limit: int) -> list[list[float]]:
        ranked = sorted(pairs, key=lambda item: item[1][0], reverse=True)[:limit]
        return [
            [material_id(code, source), round_money(values[0]), round(values[1], 2)]
            for code, values in ranked
            if values[0] > 0
        ]

    by_outlet: dict[str, dict[str, list[list[float]]]] = defaultdict(dict)
    for key, source in (("consumable", consumable_detail), ("wastage", wastage_detail)):
        grouped: dict[str, list[tuple[str, list[float]]]] = defaultdict(list)
        for (code, material), values in source["byOutlet"].items():
            grouped[code].append((material, values))
        for code, pairs in grouped.items():
            rows = top_rows(pairs, source, TOP_MATERIALS_PER_OUTLET)
            if rows:
                by_outlet[code][key] = rows

    overall = {
        key: top_rows(list(source["totals"].items()), source, TOP_MATERIALS_OVERALL)
        for key, source in (("consumable", consumable_detail), ("wastage", wastage_detail))
    }
    return catalog, dict(by_outlet), overall


def build_calendar(dates: list[str]) -> list[list[Any]]:
    """Date, ISO week and weekday index, so week grouping needs no JS date math."""
    rows: list[list[Any]] = []
    for value in dates:
        parsed = parse_date(value)
        if not parsed:
            continue
        iso = parsed.isocalendar()
        rows.append([value, iso.week, parsed.isoweekday()])
    return rows


def round_money(value: float) -> float:
    return round(value + 0.0, 2)


def build_period_alignment(sources: dict[str, dict[str, Any]]) -> dict[str, Any]:
    """Compare the covered period of every source file.

    Sales is the denominator of every rate. If the consumable or wastage export
    stops earlier than sales, the rate is divided by more days than it earns,
    which understates it. This is reported rather than silently assumed to pass.
    """
    spans = {
        name: (parse_date(quality.get("dateMin")), parse_date(quality.get("dateMax")))
        for name, quality in sources.items()
    }
    starts = [start for start, _ in spans.values() if start]
    ends = [end for _, end in spans.values() if end]
    common_start = max(starts) if starts else None
    common_end = min(ends) if ends else None
    reference_end = spans.get("sales", (None, None))[1]

    details: dict[str, Any] = {}
    for name, (start, end) in spans.items():
        details[name] = {
            "dateMin": iso_date(start),
            "dateMax": iso_date(end),
            "startLagDays": (start - common_start).days if start and common_start else None,
            "endLagDays": (reference_end - end).days if end and reference_end else None,
        }

    aligned = (
        len(starts) == len(spans)
        and len(ends) == len(spans)
        and len(set(starts)) == 1
        and len(set(ends)) == 1
    )
    lagging = sorted(
        name
        for name, detail in details.items()
        if detail["endLagDays"] not in (None, 0) and detail["endLagDays"] > 0
    )
    return {
        "aligned": aligned,
        "commonStart": iso_date(common_start),
        "commonEnd": iso_date(common_end),
        "laggingSources": lagging,
        "sources": details,
    }


def validate_payload(payload: dict[str, Any]) -> None:
    outlets = payload.get("outlets", [])
    daily = payload.get("daily", [])
    targets = payload.get("targets", [])
    if not outlets:
        raise ValueError("No outlets were produced after joining the source files.")
    if not daily:
        raise ValueError("No outlet-day records were produced from the source files.")
    codes = [outlet["code"] for outlet in outlets]
    if len(codes) != len(set(codes)):
        raise ValueError("The normalized outlet table contains duplicate outlet codes.")
    unknown_daily_codes = sorted({row["code"] for row in daily} - set(codes))
    if unknown_daily_codes:
        raise ValueError(f"Daily records contain unknown outlet codes: {', '.join(unknown_daily_codes[:20])}")
    if not any(row["sales"] for row in daily):
        raise ValueError("Overall sales are zero across the complete source period.")
    if not any(row["pnpSales"] for row in daily):
        raise ValueError("PNP Sales are zero. Confirm that Sales-Till still uses Division = FRESH PRODUCE.")
    if not any(row["consumable"] for row in daily):
        raise ValueError("No consumable movement value was found.")
    if not any(row["wastage"] for row in daily):
        raise ValueError("No wastage movement value was found.")
    if not targets or any(
        target.get(field) is None
        for target in targets
        for field in ("consumableTarget", "wastageSalesTarget", "wastagePnpTarget")
    ):
        raise ValueError("One or more benchmark target percentages are missing or invalid.")

    quality = payload.get("dataQuality", {})
    coverage = quality.get("targetSalesCoverage")
    if coverage is not None and coverage < 0.90:
        raise ValueError(
            f"Only {coverage:.1%} of sales belongs to outlets present in the zone master. "
            f"Refresh Zone-Distribution.xlsx before publishing."
        )
    reconciliation = quality.get("sales", {}).get("reconciliation", {})
    if reconciliation.get("matches") is False:
        ratio = reconciliation.get("differenceRatio") or 0
        if ratio > 0.001:
            raise ValueError(
                f"Sales detail rows total {reconciliation['detailTotal']:,.2f} but the file's own "
                f"total row says {reconciliation['reportedTotal']:,.2f} "
                f"({ratio:.3%} apart). Re-export Sales-Till.xlsx."
            )


def build_dashboard_data(input_dir: Path) -> dict[str, Any]:
    files = {
        "targets": input_dir / "Target.txt",
        "sales": input_dir / "Sales-Till.xlsx",
        "zones": input_dir / "Zone-Distribution.xlsx",
        "consumable": input_dir / "CONSUMABLE.xls",
        "wastage": input_dir / "WASTAGE.xls",
    }
    missing = [path.name for path in files.values() if not path.exists()]
    if missing:
        raise FileNotFoundError(f"Missing source files: {', '.join(missing)}")

    target_map, target_rows = read_targets(files["targets"])
    zone_map, zone_quality = read_zone_distribution(files["zones"], target_map)
    sales_daily, sales_quality = read_sales(files["sales"])
    consumable_daily, consumable_names, consumable_quality, consumable_detail = read_sap_text_export(files["consumable"])
    wastage_daily, wastage_names, wastage_quality, wastage_detail = read_sap_text_export(files["wastage"])
    material_catalog, material_by_outlet, material_leaders = build_material_view(
        consumable_detail, wastage_detail
    )

    activity_codes = {
        code for code, _ in set(sales_daily) | set(consumable_daily) | set(wastage_daily)
    }
    all_codes = sorted(set(zone_map) | activity_codes)
    outlets: list[dict[str, Any]] = []
    for code in all_codes:
        if code in zone_map:
            outlet = dict(zone_map[code])
        else:
            outlet = {
                "code": code,
                "name": first_present([consumable_names.get(code, ""), wastage_names.get(code, ""), code]),
                "regionalLeader": "Unmapped",
                "regionalLeaderFullName": "",
                "zone": "Unmapped",
                "zonalFullName": "",
                "division": "Unmapped",
                "district": "Unmapped",
                "area": "Unmapped",
                "format": "Unmapped",
                "pnpStatus": "Unmapped",
                "ownership": "Unmapped",
                "locationType": "",
                "locationClass": "",
                "populationDensity": "",
                "incomeLevel": "",
                "launchDate": None,
                "sft": None,
                "criteria": "Unmapped",
                "consumableTarget": None,
                "wastageSalesTarget": None,
                "wastagePnpTarget": None,
                "mapped": False,
            }
        outlet["hasActivity"] = code in activity_codes
        outlets.append(outlet)

    all_daily_keys = sorted(set(sales_daily) | set(consumable_daily) | set(wastage_daily), key=lambda item: (item[1], item[0]))
    daily_rows: list[dict[str, Any]] = []
    for code, date in all_daily_keys:
        sales = sales_daily.get((code, date), {})
        daily_rows.append(
            {
                "date": date.isoformat(),
                "code": code,
                "sales": round_money(sales.get("sales", 0.0)),
                "pnpSales": round_money(sales.get("pnpSales", 0.0)),
                "consumable": round_money(consumable_daily.get((code, date), 0.0)),
                "wastage": round_money(wastage_daily.get((code, date), 0.0)),
            }
        )

    consumable_totals: dict[str, float] = defaultdict(float)
    wastage_totals: dict[str, float] = defaultdict(float)
    for (code, _), value in consumable_daily.items():
        consumable_totals[code] += value
    for (code, _), value in wastage_daily.items():
        wastage_totals[code] += value

    source_max_dates = [
        parse_date(sales_quality["dateMax"]),
        parse_date(consumable_quality["dateMax"]),
        parse_date(wastage_quality["dateMax"]),
    ]
    source_max_dates = [value for value in source_max_dates if value]
    sales_total_by_code: dict[str, float] = defaultdict(float)
    for (code, _), row in sales_daily.items():
        sales_total_by_code[code] += row["sales"]
    mapped_sales = sum(value for code, value in sales_total_by_code.items() if code in zone_map)
    total_sales = sum(sales_total_by_code.values())

    data_quality = {
        "activeOutlets": len(activity_codes),
        "mappedActiveOutlets": len(activity_codes & set(zone_map)),
        "unmappedActiveOutlets": sorted(activity_codes - set(zone_map)),
        "mappedWithoutActivity": sorted(set(zone_map) - activity_codes),
        "negativeNetConsumableOutlets": sorted(
            code for code, value in consumable_totals.items() if value < 0
        ),
        "negativeNetWastageOutlets": sorted(
            code for code, value in wastage_totals.items() if value < 0
        ),
        "targetMappedOutlets": sum(
            1 for outlet in outlets if outlet["mapped"] and outlet["consumableTarget"] is not None
        ),
        "targetSalesCoverage": mapped_sales / total_sales if total_sales else None,
        "periodAlignment": build_period_alignment(
            {
                "sales": sales_quality,
                "consumable": consumable_quality,
                "wastage": wastage_quality,
            }
        ),
        "zone": zone_quality,
        "sales": sales_quality,
        "consumable": consumable_quality,
        "wastage": wastage_quality,
    }

    payload = {
        # 2: adds dataQuality.periodAlignment, sales reconciliation and SAP
        #    column-mapping diagnostics. The dashboard accepts 1 and 2.
        "schemaVersion": 3,
        "generatedAt": dt.datetime.now(dt.timezone.utc).isoformat(timespec="seconds"),
        "asOf": iso_date(min(source_max_dates) if source_max_dates else None),
        "dateRange": {
            "min": sales_quality["dateMin"],
            "max": sales_quality["dateMax"],
        },
        "metricDefinitions": {
            "pnpSales": "Sales-Till Division = FRESH PRODUCE",
            "consumableRate": "Net Consumable Value / Overall POS NSI",
            "wastageSalesRate": "Net Wastage Value / Overall POS NSI",
            "wastagePnpRate": "Net Wastage Value / FRESH PRODUCE POS NSI",
            "sapNetValue": "Movement issues less reversals, using Posting Date",
            "aggregateRate": "Sum of value divided by sum of denominator (weighted performance)",
            "outletAverage": "Simple average of valid outlet-level percentages within the selected group",
        },
        "sourceFiles": [
            {"id": "sales", "name": files["sales"].name, "rows": sales_quality["rows"], "dateMin": sales_quality["dateMin"], "dateMax": sales_quality["dateMax"]},
            {"id": "consumable", "name": files["consumable"].name, "rows": consumable_quality["rows"], "dateMin": consumable_quality["dateMin"], "dateMax": consumable_quality["dateMax"]},
            {"id": "wastage", "name": files["wastage"].name, "rows": wastage_quality["rows"], "dateMin": wastage_quality["dateMin"], "dateMax": wastage_quality["dateMax"]},
            {"id": "zones", "name": files["zones"].name, "rows": zone_quality["rows"], "dateMin": None, "dateMax": None},
            {"id": "targets", "name": files["targets"].name, "rows": len(target_rows), "dateMin": None, "dateMax": None},
        ],
        "targets": target_rows,
        "outlets": outlets,
        "daily": daily_rows,
        # [date, isoWeek, isoWeekday] for every date present in the daily rows.
        "calendar": build_calendar(sorted({row["date"] for row in daily_rows})),
        # [code, description, unit]; material rows below reference these by index.
        "materialCatalog": material_catalog,
        # code -> {consumable|wastage: [[materialId, value, quantity], ...]}
        "outletMaterials": material_by_outlet,
        # Highest-value materials across the whole selection.
        "materialLeaders": material_leaders,
        "dataQuality": data_quality,
    }
    validate_payload(payload)
    return payload


def write_dashboard_data(input_dir: Path, output_path: Path) -> dict[str, Any]:
    payload = build_dashboard_data(input_dir)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":"), allow_nan=False),
        encoding="utf-8",
    )
    return payload


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=Path("source-data"))
    parser.add_argument("--output", type=Path, default=Path("dist/data/dashboard-data.json"))
    args = parser.parse_args()
    result = write_dashboard_data(args.input, args.output)
    print(
        json.dumps(
            {
                "status": "ok",
                "outlets": len(result["outlets"]),
                "dailyRows": len(result["daily"]),
                "asOf": result["asOf"],
                "output": str(args.output),
            }
        )
    )
